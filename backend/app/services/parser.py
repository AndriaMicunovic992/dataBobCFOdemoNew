"""File parsing and schema detection service for ERP / financial data."""

from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path
from typing import Any

import polars as pl

logger = logging.getLogger(__name__)

# Polars ≥ 0.20 renamed Utf8 → String; support both versions
_Utf8: type = getattr(pl, "String", None) or pl.Utf8  # type: ignore[attr-defined]


def _normalize_utf8_view(df: pl.DataFrame) -> pl.DataFrame:
    """Cast Utf8View / LargeUtf8 columns to String before further processing.

    Newer fastexcel / Polars-Arrow may produce ``Utf8View`` columns whose
    direct casting to numeric/date types raises errors.  Normalising to the
    canonical String type first makes all downstream casts work.
    """
    casts = []
    for col_name in df.columns:
        dtype_repr = str(df[col_name].dtype)
        if "View" in dtype_repr or "LargeUtf8" in dtype_repr:
            casts.append(pl.col(col_name).cast(pl.String, strict=False).alias(col_name))
    return df.with_columns(casts) if casts else df

# ---------------------------------------------------------------------------
# Public data classes
# ---------------------------------------------------------------------------

class ParsedColumn:
    """Metadata for a single detected column."""

    __slots__ = (
        "column_name",
        "display_name",
        "data_type",
        "column_role",
        "unique_count",
        "null_count",
        "sample_values",
    )

    def __init__(
        self,
        *,
        column_name: str,
        display_name: str,
        data_type: str,
        column_role: str,
        unique_count: int,
        null_count: int,
        sample_values: list[Any],
    ) -> None:
        self.column_name = column_name
        self.display_name = display_name
        self.data_type = data_type
        self.column_role = column_role
        self.unique_count = unique_count
        self.null_count = null_count
        self.sample_values = sample_values

    def to_dict(self) -> dict[str, Any]:
        return {
            "column_name": self.column_name,
            "display_name": self.display_name,
            "data_type": self.data_type,
            "column_role": self.column_role,
            "unique_count": self.unique_count,
            "null_count": self.null_count,
            "sample_values": self.sample_values,
        }


class ParsedSheet:
    """A single table resulting from parsing (one per sheet / file)."""

    __slots__ = ("name", "data", "headers", "columns", "row_count")

    def __init__(
        self,
        *,
        name: str,
        data: pl.DataFrame,
        headers: list[str],
        columns: list[ParsedColumn],
    ) -> None:
        self.name = name
        self.data = data
        self.headers = headers
        self.columns = columns
        self.row_count = len(data)


# ---------------------------------------------------------------------------
# Column name sanitisation
# ---------------------------------------------------------------------------

_GERMAN_SUBS = str.maketrans(
    {"ä": "ae", "ö": "oe", "ü": "ue", "Ä": "Ae", "Ö": "Oe", "Ü": "Ue", "ß": "ss"}
)
_RE_NON_ALNUM = re.compile(r"[^a-z0-9]+")
_RE_LEADING_DIGIT = re.compile(r"^(\d)")


def sanitize_column_name(name: str) -> str:
    """Return a safe PostgreSQL column name (max 63 chars)."""
    # German umlauts first
    name = name.translate(_GERMAN_SUBS)
    # Normalise unicode (strip accents etc.)
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = name.lower()
    # Replace all non-alphanumeric with _
    name = _RE_NON_ALNUM.sub("_", name).strip("_")
    # Don't start with a digit
    name = _RE_LEADING_DIGIT.sub(r"col_\1", name)
    # Collapse consecutive underscores
    name = re.sub(r"_+", "_", name)
    if not name:
        name = "col"
    return name[:63]


def _deduplicate_names(names: list[str]) -> list[str]:
    """Append _2, _3 … to any duplicate sanitised names."""
    seen: dict[str, int] = {}
    result = []
    for n in names:
        if n not in seen:
            seen[n] = 0
            result.append(n)
        else:
            seen[n] += 1
            candidate = f"{n}_{seen[n] + 1}"
            while candidate in seen:
                seen[n] += 1
                candidate = f"{n}_{seen[n] + 1}"
            seen[candidate] = 0
            result.append(candidate)
    return result


# ---------------------------------------------------------------------------
# Type detection
# ---------------------------------------------------------------------------

# Date patterns common in ERP exports
_DATE_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),          # YYYY-MM-DD
    re.compile(r"^\d{4}-\d{2}$"),                 # YYYY-MM  (period)
    re.compile(r"^\d{2}\.\d{2}\.\d{4}$"),         # DD.MM.YYYY
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),           # MM/DD/YYYY
    re.compile(r"^\d{4}/\d{2}/\d{2}$"),           # YYYY/MM/DD
    re.compile(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$"),   # D.M.YY variants
    re.compile(r"^\d{1,2}\.\d{4}$"),              # M.YYYY  (German period)
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),           # DD-MM-YYYY
]
_BOOL_VALUES = frozenset({"true", "false", "yes", "no", "1", "0", "ja", "nein"})
_CURRENCY_RE = re.compile(r"^[€$£¥]?\s*-?\d[\d.,]*\s*[€$£¥]?$")


def detect_column_type(series: pl.Series) -> dict[str, Any]:
    """
    Inspect a Polars Series and return type info dict:
    {"data_type", "sample_values", "unique_count", "null_count"}.
    """
    total = len(series)
    null_count = series.null_count()
    non_null = series.drop_nulls()
    unique_count = series.n_unique()

    # Up to 1000 sample values (unique, non-null)
    sample_raw: list[Any] = non_null.unique(maintain_order=True).head(20).to_list()

    # If Polars already inferred a numeric type, trust it
    if series.dtype in (
        pl.Float32, pl.Float64, pl.Int8, pl.Int16, pl.Int32, pl.Int64,
        pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
    ):
        # Distinguish integer vs numeric
        if series.dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
            data_type = "integer"
        else:
            # Check if all finite values are whole numbers
            finite = non_null.filter(non_null.is_finite()) if non_null.dtype in (pl.Float32, pl.Float64) else non_null
            if len(finite) > 0 and (finite == finite.cast(pl.Int64).cast(pl.Float64)).all():
                data_type = "integer"
            else:
                data_type = "numeric"
        return {
            "data_type": data_type,
            "sample_values": [_safe_scalar(v) for v in sample_raw],
            "unique_count": unique_count,
            "null_count": null_count,
        }

    if series.dtype == pl.Boolean:
        return {
            "data_type": "boolean",
            "sample_values": sample_raw,
            "unique_count": unique_count,
            "null_count": null_count,
        }

    if series.dtype == pl.Date or series.dtype == pl.Datetime:
        return {
            "data_type": "date",
            "sample_values": [str(v) for v in sample_raw],
            "unique_count": unique_count,
            "null_count": null_count,
        }

    # String series: probe the actual text values
    probe = non_null.head(1000)
    probe_strs: list[str] = [str(v).strip() for v in probe.to_list()]
    n_probe = len(probe_strs)

    if n_probe == 0:
        return {
            "data_type": "text",
            "sample_values": [],
            "unique_count": unique_count,
            "null_count": null_count,
        }

    # Boolean check
    lower_vals = {v.lower() for v in probe_strs}
    if lower_vals <= _BOOL_VALUES:
        return {
            "data_type": "boolean",
            "sample_values": sample_raw,
            "unique_count": unique_count,
            "null_count": null_count,
        }

    # Date check: >80% match a date pattern
    date_hits = sum(
        1 for v in probe_strs if any(p.match(v) for p in _DATE_PATTERNS)
    )
    if date_hits / n_probe > 0.8:
        return {
            "data_type": "date",
            "sample_values": sample_raw,
            "unique_count": unique_count,
            "null_count": null_count,
        }

    # Numeric check: try casting to float after stripping currency symbols
    def _is_numeric(v: str) -> bool:
        cleaned = re.sub(r"[€$£¥\s]", "", v).replace(",", ".")
        # Handle German thousands separator (1.234,56 -> 1234.56)
        if cleaned.count(".") > 1:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            float(cleaned)
            return True
        except ValueError:
            return False

    numeric_hits = sum(1 for v in probe_strs if _is_numeric(v))
    if numeric_hits / n_probe > 0.8:
        # Integer check: all parsed values are whole numbers
        parsed = []
        for v in probe_strs:
            cleaned = re.sub(r"[€$£¥\s]", "", v).replace(",", ".")
            if cleaned.count(".") > 1:
                cleaned = cleaned.replace(".", "").replace(",", ".")
            try:
                parsed.append(float(cleaned))
            except ValueError:
                pass
        if parsed and all(f == int(f) for f in parsed):
            data_type = "integer"
        else:
            data_type = "numeric"
        return {
            "data_type": data_type,
            "sample_values": sample_raw,
            "unique_count": unique_count,
            "null_count": null_count,
        }

    return {
        "data_type": "text",
        "sample_values": sample_raw,
        "unique_count": unique_count,
        "null_count": null_count,
    }


def _safe_scalar(v: Any) -> Any:
    """Convert non-JSON-serialisable values to safe Python types."""
    if isinstance(v, float) and (v != v):  # NaN
        return None
    return v


# ---------------------------------------------------------------------------
# Date normalization
# ---------------------------------------------------------------------------

# (detection_regex, strptime_format_or_special_key, output_type)
# output_type: "iso_date" → YYYY-MM-DD, "iso_month" → YYYY-MM
_DATE_NORM_RULES: list[tuple[re.Pattern, str | None, str]] = [
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"),          None,       "iso_date"),   # already ISO
    (re.compile(r"^\d{4}-\d{2}$"),                 None,       "iso_month"),  # already ISO month
    (re.compile(r"^\d{2}\.\d{2}\.\d{4}$"),         "%d.%m.%Y", "iso_date"),   # DD.MM.YYYY
    (re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$"),     "%d.%m.%Y", "iso_date"),   # D.M.YYYY
    (re.compile(r"^\d{1,2}\.\d{4}$"),              "mm.yyyy",  "iso_month"),  # M.YYYY German period
    (re.compile(r"^\d{4}/\d{2}/\d{2}$"),           "%Y/%m/%d", "iso_date"),   # YYYY/MM/DD
    (re.compile(r"^\d{2}/\d{2}/\d{4}$"),           "%d/%m/%Y", "iso_date"),   # DD/MM/YYYY
    (re.compile(r"^\d{2}-\d{2}-\d{4}$"),           "%d-%m-%Y", "iso_date"),   # DD-MM-YYYY
]


def _detect_date_norm_rule(samples: list[str]) -> tuple[str | None, str]:
    """Return (fmt_key, output_type) for the dominant format among samples."""
    if not samples:
        return None, "iso_date"
    for pattern, fmt, output_type in _DATE_NORM_RULES:
        hits = sum(1 for s in samples if pattern.match(s))
        if hits / len(samples) >= 0.7:
            return fmt, output_type
    return None, "iso_date"


def _apply_date_norm(series: pl.Series, fmt: str, output_type: str) -> pl.Series:
    """Apply strptime normalization to a Utf8 series; preserve values that fail to parse."""
    stripped = series.str.strip_chars()

    if fmt == "mm.yyyy":
        # "01.2024" → "2024-01"  (no day component — parse manually)
        _re = re.compile(r"^(\d{1,2})\.(\d{4})$")
        result = pl.Series(
            series.name,
            [
                f"{m.group(2)}-{m.group(1).zfill(2)}"
                if v is not None and (m := _re.match(v.strip()))
                else v
                for v in series.to_list()
            ],
            dtype=_Utf8,
        )
        return result

    if output_type == "iso_month":
        parsed = stripped.str.strptime(pl.Date, format=fmt, strict=False).dt.strftime("%Y-%m")
    else:
        parsed = stripped.str.strptime(pl.Date, format=fmt, strict=False).dt.strftime("%Y-%m-%d")

    # Where parsing failed (null result), keep original stripped value
    result = pl.Series(
        series.name,
        [p if p is not None else s for p, s in zip(parsed.to_list(), stripped.to_list())],
        dtype=_Utf8,
    )
    return result


def normalize_time_series(series: pl.Series) -> tuple[pl.Series, str]:
    """
    Normalize a time-role Series to ISO Utf8 text.

    Returns (normalized_series, output_type) where output_type is
    'iso_date' (YYYY-MM-DD) or 'iso_month' (YYYY-MM).
    """
    # Polars Date → ISO string
    if series.dtype == pl.Date:
        return series.cast(_Utf8, strict=False), "iso_date"
    # Polars Datetime → ISO date string
    if series.dtype in (pl.Datetime,) or str(series.dtype).startswith("Datetime"):
        return series.dt.strftime("%Y-%m-%d"), "iso_date"

    # Normalise Utf8View before string operations
    if "View" in str(series.dtype) or "LargeUtf8" in str(series.dtype):
        series = series.cast(pl.String, strict=False)

    # String: detect format from samples
    str_series = series.cast(_Utf8, strict=False)
    samples = [s.strip() for s in str_series.drop_nulls().head(100).to_list() if s]
    if not samples:
        return str_series, "iso_date"

    fmt, output_type = _detect_date_norm_rule(samples)
    if fmt is None:
        # Already ISO or unrecognized → return as-is
        return str_series, output_type

    try:
        return _apply_date_norm(str_series, fmt, output_type), output_type
    except Exception as exc:
        logger.warning("Date normalization failed for %r: %s", series.name, exc)
        return str_series, output_type


# ---------------------------------------------------------------------------
# Role detection
# ---------------------------------------------------------------------------

_TIME_WORDS = re.compile(r"period|date|monat|month|year|jahr|datum|periode|quartal|quarter|woche|week", re.I)
_KEY_SUFFIX = re.compile(r"(_id|_code|_nr|_number|_num|konto|account|kostenstelle|cost.?center)$", re.I)
_MEASURE_WORDS = re.compile(r"amount|betrag|wert|value|quantity|menge|total|summe|sum|umsatz|revenue|cost|kosten|aufwand|expense|gewinn|profit|verlust|loss|saldo|balance", re.I)
_IGNORE_WORDS = re.compile(r"row.?count|entry.?count|zeilen.?anzahl|leer|empty", re.I)


def detect_column_role(column_name: str, type_info: dict[str, Any], total_rows: int) -> str:
    """Infer column role from name patterns and detected type."""
    dtype = type_info["data_type"]
    unique_count = type_info.get("unique_count", 0)
    null_count = type_info.get("null_count", 0)
    null_ratio = null_count / total_rows if total_rows else 0

    # Ignore: mostly null or constant values
    if null_ratio > 0.5:
        return "ignore"
    if unique_count == 1:
        return "ignore"
    if _IGNORE_WORDS.search(column_name):
        return "ignore"

    # Time dimension
    if _TIME_WORDS.search(column_name) and dtype in ("date", "text", "integer"):
        return "time"
    # Pure date columns are time regardless of name
    if dtype == "date":
        return "time"

    # Key / dimension ID
    if _KEY_SUFFIX.search(column_name):
        return "key"
    # Low-cardinality integer that looks like an account code (3-6 digits)
    if dtype == "integer" and unique_count and 2 <= unique_count <= 500:
        samples = type_info.get("sample_values", [])
        if samples and all(re.match(r"^\d{3,6}$", str(s)) for s in samples[:5]):
            return "key"

    # Measure (numeric but not a key / count)
    if dtype in ("numeric", "integer") and _MEASURE_WORDS.search(column_name):
        return "measure"
    # Fallback: numeric with high cardinality is likely a measure
    if dtype == "numeric":
        return "measure"

    # Attribute default for text
    return "attribute"


# ---------------------------------------------------------------------------
# Relationship detection
# ---------------------------------------------------------------------------

def detect_relationships(tables: dict[str, pl.DataFrame]) -> list[dict[str, Any]]:
    """
    Find likely join relationships between table pairs by checking value-set
    overlap on same-named columns (especially those ending in _id/_code/_nr).
    """
    relationships: list[dict[str, Any]] = []
    table_names = list(tables.keys())

    for i, src_name in enumerate(table_names):
        for tgt_name in table_names[i + 1 :]:
            src_df = tables[src_name]
            tgt_df = tables[tgt_name]
            src_cols = set(src_df.columns)
            tgt_cols = set(tgt_df.columns)
            # Only consider columns whose sanitised names match
            common = src_cols & tgt_cols
            for col in common:
                try:
                    src_vals = set(src_df[col].drop_nulls().unique().to_list())
                    tgt_vals = set(tgt_df[col].drop_nulls().unique().to_list())
                    if not src_vals or not tgt_vals:
                        continue
                    overlap = src_vals & tgt_vals
                    coverage = int(len(overlap) / len(src_vals) * 100)
                    if coverage >= 50:
                        relationships.append(
                            {
                                "source": src_name,
                                "target": tgt_name,
                                "source_col": col,
                                "target_col": col,
                                "coverage": coverage,
                                "overlap": len(overlap),
                            }
                        )
                        logger.debug(
                            "Relationship %s.%s -> %s.%s coverage=%d%%",
                            src_name, col, tgt_name, col, coverage,
                        )
                except Exception as exc:
                    logger.warning("Relationship check failed for %s.%s: %s", src_name, col, exc)

    return relationships


# ---------------------------------------------------------------------------
# File reading helpers
# ---------------------------------------------------------------------------

def _read_excel(file_path: Path) -> dict[str, pl.DataFrame]:
    """Read all sheets from an Excel file. Falls back to openpyxl on failure."""
    sheets: dict[str, pl.DataFrame] = {}
    try:
        import fastexcel  # noqa: F401 — confirms availability
        raw = pl.read_excel(str(file_path), sheet_id=0, engine="calamine")
        # sheet_id=0 returns all sheets as dict when engine=calamine
        if isinstance(raw, dict):
            sheets = raw
        else:
            sheets = {file_path.stem: raw}
        # Normalise Utf8View columns produced by newer fastexcel/Arrow
        sheets = {name: _normalize_utf8_view(df) for name, df in sheets.items()}
        logger.info("Read %d sheet(s) from %s via calamine", len(sheets), file_path.name)
    except Exception as exc:
        logger.warning("calamine failed (%s), falling back to openpyxl", exc)
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
                data_rows = [list(r) for r in rows[1:]]
                if not data_rows:
                    sheets[sheet_name] = pl.DataFrame({h: [] for h in headers})
                    continue
                # Build column-oriented dict to avoid dtype issues
                col_data: dict[str, list] = {h: [] for h in headers}
                for row in data_rows:
                    for h, val in zip(headers, row):
                        col_data[h].append(val)
                sheets[sheet_name] = pl.DataFrame(col_data, infer_schema_length=1000)
            logger.info("Read %d sheet(s) from %s via openpyxl", len(sheets), file_path.name)
        except Exception as exc2:
            raise RuntimeError(f"Could not parse Excel file {file_path.name}: {exc2}") from exc2
    return sheets


def _read_csv(file_path: Path) -> dict[str, pl.DataFrame]:
    """Read a CSV or TSV file, auto-detecting separator."""
    suffix = file_path.suffix.lower()
    if suffix == ".tsv":
        sep = "\t"
    else:
        # Sniff separator from first line
        try:
            first_line = file_path.read_text(encoding="utf-8-sig", errors="replace").split("\n")[0]
        except Exception:
            first_line = ""
        counts = {s: first_line.count(s) for s in (",", ";", "\t", "|")}
        sep = max(counts, key=counts.get)  # type: ignore[arg-type]
        if counts[sep] == 0:
            sep = ","  # default
        logger.debug("Detected CSV separator %r in %s", sep, file_path.name)

    try:
        df = pl.read_csv(
            str(file_path),
            separator=sep,
            infer_schema_length=1000,
            try_parse_dates=True,
            encoding="utf8-lossy",
            null_values=["", "NA", "N/A", "#N/A", "null", "NULL", "-"],
        )
        df = _normalize_utf8_view(df)
        logger.info("Read CSV %s: %d rows × %d cols", file_path.name, len(df), len(df.columns))
        return {file_path.stem: df}
    except Exception as exc:
        raise RuntimeError(f"Could not parse CSV file {file_path.name}: {exc}") from exc


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_file(file_path: str) -> dict[str, Any]:
    """
    Parse an uploaded file and return full schema analysis.

    Returns::

        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "data": <polars.DataFrame>,
                    "headers": ["Original Col Name", ...],
                    "columns": [ParsedColumn, ...],
                    "row_count": 1234,
                }
            ]
        }
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    logger.info("Parsing file %s (suffix=%s)", path.name, suffix)

    try:
        if suffix in (".xlsx", ".xls", ".xlsm", ".xlsb"):
            raw_sheets = _read_excel(path)
        elif suffix in (".csv", ".tsv", ".txt"):
            raw_sheets = _read_csv(path)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")
    except Exception as exc:
        logger.error("File read failed for %s: %s", path.name, exc)
        raise

    parsed_sheets: list[dict[str, Any]] = []

    for sheet_name, df in raw_sheets.items():
        if df.is_empty():
            logger.warning("Sheet %s is empty, skipping", sheet_name)
            continue

        logger.info("Analysing sheet %r: %d rows × %d cols", sheet_name, len(df), len(df.columns))

        original_headers = list(df.columns)
        sanitised = [sanitize_column_name(h) for h in original_headers]
        sanitised = _deduplicate_names(sanitised)

        # Rename dataframe to sanitised names
        df = df.rename(dict(zip(original_headers, sanitised)))

        total_rows = len(df)
        parsed_columns: list[ParsedColumn] = []

        for orig_header, san_name in zip(original_headers, sanitised):
            try:
                series = df[san_name]
                type_info = detect_column_type(series)
                role = detect_column_role(orig_header, type_info, total_rows)
                col = ParsedColumn(
                    column_name=san_name,
                    display_name=orig_header,
                    data_type=type_info["data_type"],
                    column_role=role,
                    unique_count=type_info["unique_count"],
                    null_count=type_info["null_count"],
                    sample_values=type_info["sample_values"],
                )
                parsed_columns.append(col)
                logger.debug(
                    "  Col %r -> %r type=%s role=%s unique=%d",
                    orig_header, san_name, type_info["data_type"], role, type_info["unique_count"],
                )
            except Exception as exc:
                logger.error("Error analysing column %r in sheet %r: %s", orig_header, sheet_name, exc)
                # Still include column with fallback type
                parsed_columns.append(
                    ParsedColumn(
                        column_name=san_name,
                        display_name=orig_header,
                        data_type="text",
                        column_role="attribute",
                        unique_count=0,
                        null_count=0,
                        sample_values=[],
                    )
                )

        # Normalize time-role columns to ISO Utf8 text so calendar joins work
        norm_casts: list[pl.Series] = []
        for col in parsed_columns:
            if col.column_role != "time":
                continue
            try:
                normalized, _iso_type = normalize_time_series(df[col.column_name])
                norm_casts.append(normalized.alias(col.column_name))
                col.data_type = "text"  # store as text so PG type matches calendar
                col.sample_values = [
                    str(v) for v in
                    normalized.drop_nulls().unique(maintain_order=True).head(20).to_list()
                    if v is not None
                ]
            except Exception as exc:
                logger.warning("Date normalization failed for col %r: %s", col.column_name, exc)
        if norm_casts:
            df = df.with_columns(norm_casts)

        parsed_sheets.append(
            {
                "name": sheet_name,
                "data": df,
                "headers": original_headers,
                "columns": parsed_columns,
                "row_count": total_rows,
            }
        )

    logger.info("Parsed %d sheet(s) from %s", len(parsed_sheets), path.name)
    return {"sheets": parsed_sheets}
